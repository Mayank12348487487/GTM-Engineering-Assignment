import openpyxl
import sqlite3
import os
import re

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "screening.db")
XLSX_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "candidate_dataset (1).xlsx")

def load_real_data():
    if not os.path.exists(XLSX_PATH):
        print(f"Dataset not found at {XLSX_PATH}")
        return
        
    print(f"Loading real data from {XLSX_PATH} into DB...")
    
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    
    # 1. Read Test Results Sheet
    test_results = {}
    sheet_test = wb["Test Result"]
    for row in sheet_test.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            name = str(row[1]).strip()
            test_results[name] = {
                "email": str(row[2]).strip(),
                "la_score": float(row[6]) if row[6] is not None else 0.0,
                "code_score": float(row[7]) if row[7] is not None else 0.0
            }
            
    # 2. Connect to Database & Clear tables
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    print("Clearing existing candidates, interviews, and email logs...")
    cursor.execute("DELETE FROM candidates")
    cursor.execute("DELETE FROM email_logs")
    cursor.execute("DELETE FROM interviews")
    
    # 3. Read Responses and insert candidates
    sheet_resp = wb["Response"]
    inserted_count = 0
    
    for row in sheet_resp.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            s_no = row[0]
            name = str(row[1]).strip()
            base_email = str(row[2]).strip()
            college = str(row[3]).strip() if row[3] is not None else ""
            branch = str(row[4]).strip() if row[4] is not None else ""
            cgpa_val = float(row[5]) if row[5] is not None else 0.0
            best_project = str(row[6]).strip() if row[6] is not None else ""
            research = str(row[7]).strip() if row[7] is not None else ""
            github = str(row[8]).strip() if row[8] is not None else ""
            resume = str(row[9]).strip() if row[9] is not None else ""
            
            # Format subaddressed unique email to bypass UNIQUE constraint
            email_match = re.match(r"^([^@]+)(@.+)$", base_email)
            if email_match:
                prefix, suffix = email_match.groups()
                unique_email = f"{prefix}+student{s_no}{suffix}"
            else:
                unique_email = f"student{s_no}@example.com"
                
            # Check if this student has test results
            if name in test_results:
                test_la = test_results[name]["la_score"]
                test_code = test_results[name]["code_score"]
                test_status = "Completed"
                status = "Test Completed"
            else:
                test_la = -1.0
                test_code = -1.0
                test_status = "Not Sent"
                status = "Applied"
                
            cursor.execute("""
                INSERT INTO candidates 
                (name, email, college, branch, cgpa, best_ai_project, research_work, github_profile, resume_link, status, test_la_score, test_code_score, test_status) 
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (name, unique_email, college, branch, cgpa_val, best_project, research, github, resume, status, test_la, test_code, test_status))
            inserted_count += 1
            
    conn.commit()
    conn.close()
    print(f"Data loading complete. Successfully imported {inserted_count} candidates.")

if __name__ == "__main__":
    load_real_data()
